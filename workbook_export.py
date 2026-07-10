"""
Instructional Workbook Export
==============================
Generates a leader-facing, multi-tab Excel workbook from the canonical
per-student DataFrame that any ingest module (iReady, IXL, MAP) already
produces -- plus optional demographic/prior-year columns when the upload
has them.

Design goals (improvements over hand-built BOY results workbooks):
  - ONE roster, filterable, instead of the same students duplicated across
    subject/GAP tabs. GAP is a computed flag, not a separate copy.
  - LIVE formulas for every aggregate (COUNTIFS), so a corrected row on the
    Roster flows through Summary automatically. No stale static counts.
  - Headers on a predictable row with freeze panes + autofilter; percents
    formatted as percents.
  - The tier/priority engine (same logic as the standalone template) so the
    workbook is decision-ready, not just descriptive.
  - A Projected NSPF tab connecting student-level data to the accountability
    outcome -- the link hand-built workbooks usually lack.
  - A Data Quality tab replacing NotTested tabs, computed not pasted.

Privacy: this module builds the workbook IN MEMORY and returns bytes. The
app serves it via st.download_button; nothing is written server-side, so the
app's "nothing saved to disk" promise holds. The generated file itself
contains student-level data -- the About tab tells the leader to treat it
like any student record.

Expected DataFrame columns (missing optional columns are simply skipped):
  required: student_id, grade, subject ('ELA'/'MATH'), probable_level (1-4)
  optional: first_name, last_name, percentile, growth_pct, met_growth (bool), prior_level (1-4),
            race, iep (Y/blank), ell (Y/blank), target_group (school labels
            like Blue/Purple), scale_score, points_to_level_up
"""

from __future__ import annotations
import io
from datetime import datetime
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference

F = "Arial"
BLUE = Font(name=F, size=10, color="0000FF")
BLACK = Font(name=F, size=10)
BOLD = Font(name=F, size=10, bold=True)
ITAL = Font(name=F, size=9, italic=True, color="555555")
WHT = Font(name=F, size=12, bold=True, color="FFFFFF")
HFILL = PatternFill("solid", start_color="1F3864")
SFILL = PatternFill("solid", start_color="D9E2F2")
YEL = PatternFill("solid", start_color="FFFF00")
THIN = Border(bottom=Side(style="thin", color="BBBBBB"))
CTR = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")
PCT = "0.0%"

OPTIONAL_COLS = ["first_name", "last_name", "percentile", "growth_pct", "met_growth", "prior_level", "race",
                 "iep", "ell", "target_group", "scale_score", "points_to_level_up"]


def _norm(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in OPTIONAL_COLS:
        if col not in df.columns:
            df[col] = pd.NA
    df["subject"] = df["subject"].astype(str).str.upper().str.strip()
    for col in ["probable_level", "percentile", "growth_pct", "prior_level",
                "scale_score", "points_to_level_up", "grade"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df




def _add_charts_tab(wb, df, L, start, end, S_PROF):
    """Charts tab: live COUNTIFS tables (shown, not hidden) feeding Excel-native
    charts. Because the tables reference the Roster, the charts update when the
    Roster is corrected -- same live-data philosophy as the Summary."""
    ch = wb.create_sheet("Charts")
    ch["A1"] = "CHARTS — every chart is fed by the labeled table beside it (live from the Roster)"
    ch["A1"].font = WHT
    for i in range(1, 13):
        ch.cell(row=1, column=i).fill = HFILL
    ch.column_dimensions["A"].width = 26
    for col in "BCDE":
        ch.column_dimensions[col].width = 11

    SUBJ = f"Roster!${L['subject']}${start}:${L['subject']}${end}"
    LVL = f"Roster!${L['level']}${start}:${L['level']}${end}"
    GRD = f"Roster!${L['grade']}${start}:${L['grade']}${end}"
    TIER = f"Roster!${L['tier']}${start}:${L['tier']}${end}"
    grades = sorted({int(g) for g in df["grade"].dropna().unique()})

    # ---- Table/Chart 1: % projected proficient by grade ----
    r0 = 3
    ch.cell(row=r0, column=1, value="Projected % proficient").font = BOLD
    for i, h in enumerate(["", "ELA", "MATH"], 1):
        c = ch.cell(row=r0 + 1, column=i, value=h); c.font = BOLD; c.fill = SFILL; c.alignment = CTR
    rr = r0 + 2
    for label, g in [("Pooled", None)] + [(f"Grade {g}", g) for g in grades]:
        ch.cell(row=rr, column=1, value=label).font = BLACK
        for col, subj in ((2, "ELA"), (3, "MATH")):
            gbit = f",{GRD},{g}" if g is not None else ""
            f_ = (f'=IF(COUNTIFS({SUBJ},"{subj}",{LVL},"<>"&""{gbit})=0,0,'
                  f'COUNTIFS({SUBJ},"{subj}",{LVL},">="&{S_PROF}{gbit})/'
                  f'COUNTIFS({SUBJ},"{subj}",{LVL},"<>"&""{gbit}))')
            c = ch.cell(row=rr, column=col, value=f_); c.number_format = PCT; c.alignment = CTR
        rr += 1
    t1_end = rr - 1
    bar1 = BarChart(); bar1.type = "col"; bar1.style = 10
    bar1.title = "Projected % proficient by grade"
    bar1.y_axis.numFmt = "0%"; bar1.y_axis.title = "% at/above proficient"
    data = Reference(ch, min_col=2, max_col=3, min_row=r0 + 1, max_row=t1_end)
    cats = Reference(ch, min_col=1, min_row=r0 + 2, max_row=t1_end)
    bar1.add_data(data, titles_from_data=True); bar1.set_categories(cats)
    bar1.width = 16; bar1.height = 9
    ch.add_chart(bar1, "G3")

    # ---- Table/Chart 2: projected level distribution (stacked) ----
    r0 = t1_end + 3
    ch.cell(row=r0, column=1, value="Students by projected level").font = BOLD
    for i, h in enumerate(["", "Level 1", "Level 2", "Level 3", "Level 4"], 1):
        c = ch.cell(row=r0 + 1, column=i, value=h); c.font = BOLD; c.fill = SFILL; c.alignment = CTR
    rr = r0 + 2
    for subj in ("ELA", "MATH"):
        ch.cell(row=rr, column=1, value=subj).font = BLACK
        for lv in (1, 2, 3, 4):
            c = ch.cell(row=rr, column=1 + lv,
                        value=f'=COUNTIFS({SUBJ},"{subj}",{LVL},{lv})')
            c.alignment = CTR
        rr += 1
    t2_end = rr - 1
    bar2 = BarChart(); bar2.type = "col"; bar2.grouping = "stacked"; bar2.overlap = 100
    bar2.style = 11; bar2.title = "Projected level distribution"
    bar2.y_axis.title = "Students"
    data = Reference(ch, min_col=2, max_col=5, min_row=r0 + 1, max_row=t2_end)
    cats = Reference(ch, min_col=1, min_row=r0 + 2, max_row=t2_end)
    bar2.add_data(data, titles_from_data=True); bar2.set_categories(cats)
    bar2.width = 16; bar2.height = 9
    ch.add_chart(bar2, "G22")

    # ---- Table/Chart 3: tiers by subject (horizontal, long labels) ----
    r0 = t2_end + 3
    ch.cell(row=r0, column=1, value="Students by tier").font = BOLD
    for i, h in enumerate(["", "ELA", "MATH"], 1):
        c = ch.cell(row=r0 + 1, column=i, value=h); c.font = BOLD; c.fill = SFILL; c.alignment = CTR
    rr = r0 + 2
    for t in ["1 - Intensive Intervention", "2 - Targeted Push (Bubble)", "2 - Strategic Support",
              "3 - Core / Maintain", "4 - Extension / Enrich"]:
        ch.cell(row=rr, column=1, value=t).font = BLACK
        ch.cell(row=rr, column=2, value=f'=COUNTIFS({TIER},$A{rr},{SUBJ},"ELA")').alignment = CTR
        ch.cell(row=rr, column=3, value=f'=COUNTIFS({TIER},$A{rr},{SUBJ},"MATH")').alignment = CTR
        rr += 1
    t3_end = rr - 1
    bar3 = BarChart(); bar3.type = "bar"; bar3.style = 12
    bar3.title = "Instructional tiers by subject"
    bar3.x_axis.title = "Students"
    data = Reference(ch, min_col=2, max_col=3, min_row=r0 + 1, max_row=t3_end)
    cats = Reference(ch, min_col=1, min_row=r0 + 2, max_row=t3_end)
    bar3.add_data(data, titles_from_data=True); bar3.set_categories(cats)
    bar3.width = 16; bar3.height = 10
    ch.add_chart(bar3, "G41")

    # ---- Reading guide ----
    r0 = t3_end + 2
    notes = [
        "How to read these:",
        "Chart 1 is the headline: where each grade stands against the proficiency bar. Grades far below pooled are where staffing follows.",
        "Chart 2 shows composition: a tall Level 2 block is opportunity (bubble students one band from proficient); a tall Level 1 block is an intervention-capacity question.",
        "Chart 3 is the staffing picture: how many students each tier must serve, per subject. If Tier 1 exceeds intervention capacity, the plan, not the data, is the problem.",
    ]
    for note in notes:
        c = ch.cell(row=r0, column=1, value=note)
        c.font = BOLD if note.endswith(":") else ITAL
        c.alignment = WRAP
        r0 += 1


def build_instructional_workbook(
    df: pd.DataFrame,
    school_name: str = "School",
    platform: str = "interim assessment",
    term: str = "",
    nspf_result=None,           # nspf_engine.Result, optional
    nspf_rates: Optional[dict] = None,
    level_key: str = "Middle",
    low_growth_cut: float = 35,
    bubble_pctile_cut: float = 45,
) -> bytes:
    """Return xlsx bytes. See module docstring for the expected columns."""
    df = _norm(df)
    has = {c: df[c].notna().any() for c in OPTIONAL_COLS}
    n_rows = len(df)
    start, end = 6, 5 + max(n_rows, 1)

    wb = Workbook()

    # ---------------- About ----------------
    ab = wb.active
    ab.title = "About"
    ab.column_dimensions["A"].width = 108
    lines = [
        ("T", f"{school_name} — Instructional Decision Workbook"),
        ("P", f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} from a {platform} upload"
              + (f" ({term})" if term else "") + f" · {n_rows} student-subject rows."),
        ("H", "How to use this"),
        ("P", "ROSTER is the working surface: every student, both subjects, with computed tier, "
              "flags, and a 0-9 priority. Use the filter arrows (e.g. Subject=MATH, Tier=1, sort "
              "by Priority) to build any group. SUMMARY holds live aggregates (they update if you "
              "correct a Roster input). SETTINGS thresholds drive every flag. PROJECTED NSPF ties "
              "this roster to the school's accountability picture. DATA QUALITY shows what was "
              "missing from the upload."),
        ("H", "Tier definitions"),
        ("P", "1 Intensive = projected Level 1.  2 Targeted Push (Bubble) = Level 2 with strong "
              "percentile — the band-edge group closest to proficient.  2 Strategic = other Level 2. "
              "3 Core = projected proficient.  4 Extension = projected Level 4.  GAP = below "
              "proficient on the prior-year state test (their growth feeds NSPF's Closing "
              "Opportunity Gaps).  Growth Alert = low growth percentile or missed growth target."),
        ("H", "This is a projection, not an official score"),
        ("P", "Tiers and rates here derive from an interim assessment, not state results. They are "
              "planning inputs. Nothing in this workbook is suitable for public, board-external, "
              "or authorizer-facing reporting."),
        ("H", "Data privacy"),
        ("P", "This file contains student-level records. Store it where your school stores student "
              "records; share only with staff who have a legitimate educational interest; never "
              "email unencrypted or paste contents into outside tools. The SUMMARY tab alone "
              "(aggregates, no names) is the shareable layer."),
    ]
    r = 1
    for kind, text in lines:
        c = ab.cell(row=r, column=1, value=text)
        if kind == "T":
            c.font = WHT; c.fill = HFILL
        elif kind == "H":
            c.font = BOLD; c.fill = SFILL
        else:
            c.font = BLACK; c.alignment = WRAP
            ab.row_dimensions[r].height = max(15, 13 * (len(text) // 100 + 1))
        r += 1

    # ---------------- Settings ----------------
    se = wb.create_sheet("Settings")
    se.column_dimensions["A"].width = 52
    se.column_dimensions["B"].width = 10
    se.column_dimensions["C"].width = 66
    se["A1"] = "Settings — thresholds that drive every flag (blue = edit)"
    se["A1"].font = WHT
    for col in "ABC":
        se[f"{col}1"].fill = HFILL
    for i, (label, val, note, hot) in enumerate([
        ("Proficient at projected level >=", 3, "Level 3+ counts as projected proficient.", False),
        ("Low-growth alert: growth percentile below", low_growth_cut, "Students under this get a GROWTH ALERT.", True),
        ("Bubble: Level 2 AND percentile at/above", bubble_pctile_cut, "The band-edge push group.", True),
    ], start=3):
        se.cell(row=i, column=1, value=label).font = BLACK
        v = se.cell(row=i, column=2, value=val); v.font = BLUE; v.alignment = CTR
        if hot:
            v.fill = YEL
        n = se.cell(row=i, column=3, value=note); n.font = ITAL; n.alignment = WRAP
    S_PROF, S_LOWG, S_BUB = "Settings!$B$3", "Settings!$B$4", "Settings!$B$5"

    # ---------------- Roster ----------------
    ro = wb.create_sheet("Roster")
    cols = [("Student ID", 12), ("Student", 22), ("Grade", 7), ("Subject", 9)]
    if has["race"]:
        cols.append(("Race/Ethnicity", 20))
    if has["iep"]:
        cols.append(("IEP", 6))
    if has["ell"]:
        cols.append(("ELL", 6))
    if has["target_group"]:
        cols.append(("Target Group", 12))
    if has["prior_level"]:
        cols.append(("Prior Yr Level", 10))
    cols.append(("Proj. Level", 9))
    if has["percentile"]:
        cols.append(("Percentile", 9))
    if has["growth_pct"]:
        cols.append(("Growth %ile", 9))
    if has["met_growth"]:
        cols.append(("Met Growth", 9))
    if has["points_to_level_up"]:
        cols.append(("Pts to Level Up", 11))
    computed_start = len(cols) + 1
    cols += [("Tier", 26), ("Bubble", 8), ("GAP", 7), ("Growth Alert", 10),
             ("Priority", 8), ("Suggested Focus", 56)]

    ro["A1"] = ("ROSTER — one filterable table (GAP is the flag column, not a separate tab). "
                "Data columns from the upload; Tier onward computed.")
    ro["A1"].font = WHT
    for i in range(1, len(cols) + 1):
        ro.cell(row=1, column=i).fill = HFILL
    ro["A3"] = "Filter example: Subject=MATH + Tier starts with 1, sorted by Priority = Monday's intervention list."
    ro["A3"].font = ITAL
    for i, (h, w) in enumerate(cols, 1):
        c = ro.cell(row=5, column=i, value=h)
        c.font = BOLD; c.fill = SFILL; c.border = THIN
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ro.column_dimensions[get_column_letter(i)].width = w
    ro.freeze_panes = "E6"

    # letters for referenced columns
    L = {}
    ci = 1
    for name in ["id", "student", "grade", "subject"]:
        L[name] = get_column_letter(ci); ci += 1
    for opt, key in [("race", "race"), ("iep", "iep"), ("ell", "ell"),
                     ("target_group", "target"), ("prior_level", "prior")]:
        if has[opt]:
            L[key] = get_column_letter(ci); ci += 1
    L["level"] = get_column_letter(ci); ci += 1
    for opt, key in [("percentile", "pctile"), ("growth_pct", "gpct"),
                     ("met_growth", "metg"), ("points_to_level_up", "ptsup")]:
        if has[opt]:
            L[key] = get_column_letter(ci); ci += 1
    L["tier"], L["bub"], L["gap"], L["alert"], L["pri"], L["focus"] = (
        get_column_letter(computed_start + k) for k in range(6))

    for j, (_, row) in enumerate(df.iterrows()):
        r = start + j
        def put(key, val, center=True):
            c = ro.cell(row=r, column=ro[f"{L[key]}1"].column, value=val)
            c.font = BLACK
            if center:
                c.alignment = CTR
            return c
        put("id", row["student_id"], center=False)
        nm = f"{row['first_name'] if pd.notna(row['first_name']) else ''} "\
             f"{row['last_name'] if pd.notna(row['last_name']) else ''}".strip()
        put("student", nm if nm else row["student_id"], center=False)
        put("grade", row["grade"] if pd.notna(row["grade"]) else None)
        put("subject", row["subject"])
        if has["race"]:
            put("race", row["race"] if pd.notna(row["race"]) else None, center=False)
        if has["iep"]:
            put("iep", row["iep"] if pd.notna(row["iep"]) else None)
        if has["ell"]:
            put("ell", row["ell"] if pd.notna(row["ell"]) else None)
        if has["target_group"]:
            put("target", row["target_group"] if pd.notna(row["target_group"]) else None)
        if has["prior_level"]:
            put("prior", row["prior_level"] if pd.notna(row["prior_level"]) else None)
        put("level", row["probable_level"] if pd.notna(row["probable_level"]) else None)
        if has["percentile"]:
            put("pctile", row["percentile"] if pd.notna(row["percentile"]) else None)
        if has["growth_pct"]:
            put("gpct", row["growth_pct"] if pd.notna(row["growth_pct"]) else None)
        if has["met_growth"]:
            mg = row["met_growth"]
            put("metg", ("Y" if mg else "N") if pd.notna(mg) else None)
        if has["points_to_level_up"]:
            put("ptsup", row["points_to_level_up"] if pd.notna(row["points_to_level_up"]) else None)

        lv = f"${L['level']}{r}"
        bub_expr = (f'IF(AND({lv}=2,{L["pctile"]}{r}<>"",{L["pctile"]}{r}>={S_BUB}),"Yes","No")'
                    if has["percentile"] else '"No"')
        ro[f"{L['tier']}{r}"] = (
            f'=IF({lv}="","",IF({lv}<=1,"1 - Intensive Intervention",'
            f'IF({lv}=2,IF({L["bub"]}{r}="Yes","2 - Targeted Push (Bubble)","2 - Strategic Support"),'
            f'IF({lv}={S_PROF},"3 - Core / Maintain","4 - Extension / Enrich"))))')
        ro[f"{L['bub']}{r}"] = f'=IF({lv}="","",{bub_expr})'
        if has["prior_level"]:
            ro[f"{L['gap']}{r}"] = (f'=IF({lv}="","",IF({L["prior"]}{r}="","?",'
                                    f'IF({L["prior"]}{r}<{S_PROF},"Yes","No")))')
        else:
            ro[f"{L['gap']}{r}"] = f'=IF({lv}="","","?")'
        alert_parts = []
        if has["growth_pct"]:
            alert_parts.append(f'AND({L["gpct"]}{r}<>"",{L["gpct"]}{r}<{S_LOWG})')
        if has["met_growth"]:
            alert_parts.append(f'{L["metg"]}{r}="N"')
        if alert_parts:
            ro[f"{L['alert']}{r}"] = f'=IF({lv}="","",IF(OR({",".join(alert_parts)}),"Yes","No"))'
        else:
            ro[f"{L['alert']}{r}"] = f'=IF({lv}="","","?")'
        ro[f"{L['pri']}{r}"] = (
            f'=IF({lv}="","",({lv}=1)*3+({lv}=2)*2+({L["bub"]}{r}="Yes")*2'
            f'+({L["gap"]}{r}="Yes")*2+({L["alert"]}{r}="Yes")*1)')
        up_bit = (f'&IF({L["ptsup"]}{r}<>"",";  "&{L["ptsup"]}{r}&" scale pts to next level","")'
                  if has["points_to_level_up"] else "")
        ro[f"{L['focus']}{r}"] = (
            f'=IF({L["tier"]}{r}="","",'
            f'IF(LEFT({L["tier"]}{r},1)="1","Daily small-group intervention; diagnose skill gaps",'
            f'IF({L["tier"]}{r}="2 - Targeted Push (Bubble)","Band-edge: targeted standards push",'
            f'IF(LEFT({L["tier"]}{r},1)="2","Strategic small-group 2-3x/week",'
            f'IF(LEFT({L["tier"]}{r},1)="3","Strong core; recheck next window","Extend and deepen"))))'
            f'&IF({L["alert"]}{r}="Yes","; + growth intervention","")'
            f'&IF({L["gap"]}{r}="Yes","; GAP: growth earns NSPF gap points","")'
            f'{up_bit})')
        for key in ["tier", "bub", "gap", "alert", "pri"]:
            ro[f"{L[key]}{r}"].alignment = CTR
        ro[f"{L['focus']}{r}"].alignment = WRAP

    ro.auto_filter.ref = f"A5:{get_column_letter(len(cols))}{end}"
    ro.conditional_formatting.add(
        f"{L['pri']}{start}:{L['pri']}{end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       mid_type="num", mid_value=4, mid_color="FFE699",
                       end_type="num", end_value=9, end_color="F4B183"))

    # ---------------- Summary (all live COUNTIFS) ----------------
    su = wb.create_sheet("Summary")
    su["A1"] = "SUMMARY — live aggregates (update automatically if the Roster changes). Shareable: no names."
    su["A1"].font = WHT
    for i in range(1, 8):
        su.cell(row=1, column=i).fill = HFILL
    su.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        su.column_dimensions[col].width = 11

    SUBJ = f"Roster!${L['subject']}${start}:${L['subject']}${end}"
    LVL = f"Roster!${L['level']}${start}:${L['level']}${end}"
    GRD = f"Roster!${L['grade']}${start}:${L['grade']}${end}"
    TIER = f"Roster!${L['tier']}${start}:${L['tier']}${end}"
    GAP = f"Roster!${L['gap']}${start}:${L['gap']}${end}"

    def prof_block(row0, title, extra_rng=None, extra_crit=None):
        su.cell(row=row0, column=1, value=title).font = BOLD
        hdrs = ["", "ELA n", "ELA meets", "ELA %", "MATH n", "MATH meets", "MATH %"]
        for i, h in enumerate(hdrs, 1):
            c = su.cell(row=row0 + 1, column=i, value=h); c.font = BOLD; c.fill = SFILL; c.alignment = CTR
        grades = sorted({int(g) for g in df["grade"].dropna().unique()})
        labels = [("Pooled", None)] + [(str(g), g) for g in grades]
        rr = row0 + 2
        for label, g in labels:
            su.cell(row=rr, column=1, value=label).font = BLACK
            for base, subj in ((2, "ELA"), (5, "MATH")):
                crit = [SUBJ, f'"{subj}"', LVL, '"<>"&""']
                crit_m = [SUBJ, f'"{subj}"', LVL, f'">="&{S_PROF}']
                if g is not None:
                    crit += [GRD, str(g)]; crit_m += [GRD, str(g)]
                if extra_rng:
                    crit += [extra_rng, extra_crit]; crit_m += [extra_rng, extra_crit]
                nf = f'=COUNTIFS({",".join(crit)})'
                mf = f'=COUNTIFS({",".join(crit_m)})'
                nl, ml, pl = get_column_letter(base), get_column_letter(base + 1), get_column_letter(base + 2)
                su.cell(row=rr, column=base, value=nf).alignment = CTR
                su.cell(row=rr, column=base + 1, value=mf).alignment = CTR
                p = su.cell(row=rr, column=base + 2, value=f'=IF({nl}{rr}=0,"",{ml}{rr}/{nl}{rr})')
                p.number_format = PCT; p.alignment = CTR; p.font = BOLD
            rr += 1
        return rr

    r = prof_block(3, "Projected proficiency (probable level >= proficient) by grade")
    r = prof_block(r + 2, "GAP students only (prior year below proficient)", GAP, '"Yes"') if has["prior_level"] else r
    if has["target_group"]:
        TGT = f"Roster!${L['target']}${start}:${L['target']}${end}"
        r = prof_block(r + 2, "Target-group students (school designations)", TGT, '"<>"&""')

    r += 2
    su.cell(row=r, column=1, value="Students by tier").font = BOLD
    r += 1
    for i, h in enumerate(["", "ELA", "MATH"], 1):
        c = su.cell(row=r, column=i, value=h); c.font = BOLD; c.fill = SFILL; c.alignment = CTR
    for t in ["1 - Intensive Intervention", "2 - Targeted Push (Bubble)", "2 - Strategic Support",
              "3 - Core / Maintain", "4 - Extension / Enrich"]:
        r += 1
        su.cell(row=r, column=1, value=t).font = BLACK
        su.cell(row=r, column=2, value=f'=COUNTIFS({TIER},$A{r},{SUBJ},"ELA")').alignment = CTR
        su.cell(row=r, column=3, value=f'=COUNTIFS({TIER},$A{r},{SUBJ},"MATH")').alignment = CTR

    if has["race"]:
        RACE = f"Roster!${L['race']}${start}:${L['race']}${end}"
        r += 2
        su.cell(row=r, column=1, value="Projected proficiency by race/ethnicity").font = BOLD
        r += 1
        for i, h in enumerate(["", "ELA n", "ELA %", "MATH n", "MATH %"], 1):
            c = su.cell(row=r, column=i, value=h); c.font = BOLD; c.fill = SFILL; c.alignment = CTR
        for race in sorted(df["race"].dropna().astype(str).unique()):
            r += 1
            su.cell(row=r, column=1, value=race).font = BLACK
            for base, subj in ((2, "ELA"), (4, "MATH")):
                nl = get_column_letter(base)
                su.cell(row=r, column=base,
                        value=f'=COUNTIFS({SUBJ},"{subj}",{RACE},$A{r},{LVL},"<>"&"")').alignment = CTR
                p = su.cell(row=r, column=base + 1,
                            value=(f'=IF({nl}{r}=0,"",COUNTIFS({SUBJ},"{subj}",{RACE},$A{r},'
                                   f'{LVL},">="&{S_PROF})/{nl}{r})'))
                p.number_format = PCT; p.alignment = CTR

    _add_charts_tab(wb, df, L, start, end, S_PROF)

    # ---------------- Projected NSPF ----------------
    if nspf_result is not None:
        np_ = wb.create_sheet("Projected NSPF")
        np_.column_dimensions["A"].width = 46
        np_.column_dimensions["B"].width = 14
        np_.column_dimensions["C"].width = 60
        np_["A1"] = "PROJECTED NSPF — snapshot at generation (regenerate from the app to refresh)"
        np_["A1"].font = WHT
        for col in "ABC":
            np_[f"{col}1"].fill = HFILL
        np_["A2"] = ("PROJECTED / INTERIM — not an official NDE rating. Same disclaimers as the app; "
                     "the scoring engine is exact, the input rates are interim-based.")
        np_["A2"].font = Font(name=F, size=9, bold=True, color="B43C00")
        np_["A4"] = "Projected Index"; np_["A4"].font = BOLD
        np_["B4"] = nspf_result.index; np_["B4"].font = BOLD
        np_["A5"] = "Projected Stars"; np_["A5"].font = BOLD
        np_["B5"] = "*" * nspf_result.stars; np_["B5"].font = BOLD
        np_["A6"] = "Rated (all required measures present)"; np_["B6"] = "Yes" if nspf_result.rated else "No (provisional)"
        rr = 8
        np_.cell(row=rr, column=1, value="Measure rates used").font = BOLD
        rr += 1
        for k, v in (nspf_rates or {}).items():
            np_.cell(row=rr, column=1, value=k).font = BLACK
            np_.cell(row=rr, column=2, value=v if v is not None else "not available").font = BLACK
            rr += 1
        rr += 1
        for comp, c in nspf_result.by_component.items():
            np_.cell(row=rr, column=1, value=comp).font = BLACK
            np_.cell(row=rr, column=2, value=f"{c['earned']:.1f} / {c['possible']:g}").font = BLACK
            rr += 1

    # ---------------- Data Quality ----------------
    dq = wb.create_sheet("Data Quality")
    dq.column_dimensions["A"].width = 44
    dq.column_dimensions["B"].width = 12
    dq["A1"] = "DATA QUALITY — what the upload was missing (live counts)"
    dq["A1"].font = WHT
    dq["B1"].fill = HFILL; dq["A1"].fill = HFILL
    checks = [("Rows with no projected level (not tested?)",
               f'=COUNTIFS({SUBJ},"<>"&"",{LVL},"")'),
              ("Rows total", f'=COUNTIFS({SUBJ},"<>"&"")')]
    if has["prior_level"]:
        PRI = f"Roster!${L['prior']}${start}:${L['prior']}${end}"
        checks.insert(1, ("Rows missing prior-year level (GAP flag = ?)",
                          f'=COUNTIFS({SUBJ},"<>"&"",{PRI},"")'))
    if has["growth_pct"]:
        GP = f"Roster!${L['gpct']}${start}:${L['gpct']}${end}"
        checks.insert(1, ("Rows missing growth percentile",
                          f'=COUNTIFS({SUBJ},"<>"&"",{GP},"")'))
    for i, (label, f_) in enumerate(checks, start=3):
        dq.cell(row=i, column=1, value=label).font = BLACK
        dq.cell(row=i, column=2, value=f_).alignment = CTR

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
